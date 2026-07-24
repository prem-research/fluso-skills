#!/usr/bin/env python3
"""Deterministic spreadsheet I/O for the security-questionnaire skill.

This helper does the mechanical work only. It never invents, infers, or
scores an answer -- that is the skill's job, done from cited policy sources.
The tool exists so the model reads a real questionnaire as clean rows and
writes vetted answers back into the *exact* cells, leaving column layout and
worksheet formatting untouched.

Subcommands
-----------
extract   <input.xlsx|.csv>              -> JSON: detected column map + one object per question row
template  <rows.json>                    -> JSON: a blank answer skeleton to fill in
fill      <input.xlsx|.csv> <answers.json> -> writes answers into a copy, formatting preserved

Everything runs locally. No network access is used or required.

Run it through the skill's isolated environment so openpyxl is available:

    uv run --project /fluso/user/workspace/skills/security-questionnaire \\
        python scripts/questionnaire.py extract questionnaire.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

# Fields the skill fills in. Detection maps each to at most one column.
ANSWER_FIELDS = ("response", "detail", "evidence", "status")
CONTEXT_FIELDS = ("ref", "domain", "question")
ALL_FIELDS = CONTEXT_FIELDS + ANSWER_FIELDS

# Normalized header tokens seen across CAIQ, SIG, VSAQ, HECVAT, and custom
# sheets. Matching is exact-on-normalized first, then whole-word fallback.
HEADER_SYNONYMS: dict[str, set[str]] = {
    "ref": {
        "ref", "refid", "referenceid", "id", "no", "num", "number", "item",
        "controlid", "ctrl", "qid", "questionid", "questionnumber", "cid",
    },
    "domain": {
        "domain", "category", "section", "area", "controlfamily", "family",
        "topic", "group", "controldomain", "principle",
    },
    "question": {
        "question", "control", "requirement", "description", "query", "ask",
        "prompt", "statement", "criteria", "controlspecification",
        "questiontext", "controlquestion",
    },
    "response": {
        "response", "answer", "yesno", "yn", "compliance", "compliant",
        "vendorresponse", "responseyesno", "answeryesno",
    },
    "detail": {
        "detail", "details", "comment", "comments", "note", "notes",
        "explanation", "justification", "remarks", "elaboration",
        "additionalinformation", "responsedetail", "vendorcomments",
        "notesexplanation", "commentsnotes",
    },
    "evidence": {
        "evidence", "reference", "references", "referencedocument", "citation",
        "source", "sources", "attachment", "attachments", "supportingevidence",
        "document", "artifact", "link", "evidencelink", "supportingdocumentation",
    },
    "status": {
        "status", "reviewstatus", "state", "completion", "disposition",
        "completionstatus",
    },
}

# When one header cell matches several fields by whole word (e.g.
# "Response Detail" -> response + detail), the more specific field wins.
FIELD_PRIORITY = ("detail", "evidence", "status", "ref", "domain", "response", "question")


class QError(Exception):
    """A user-facing failure with a clean message (no traceback)."""


def normalize(value: object) -> str:
    """Lowercase and drop everything but a-z0-9 so headers compare cleanly."""
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def col_letter(index0: int) -> str:
    """0-based column index -> spreadsheet letter (0 -> A, 26 -> AA)."""
    letters = ""
    n = index0 + 1
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# --------------------------------------------------------------------------- #
# Reading: build a plain grid (list of rows of string cells) from csv or xlsx.
# --------------------------------------------------------------------------- #

def read_csv_grid(path: Path) -> tuple[list[list[str]], str]:
    text = path.read_text(encoding="utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    return [[("" if c is None else str(c)) for c in row] for row in rows], delimiter


def read_xlsx_grid(path: Path, sheet: str | None) -> tuple[list[list[str]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise QError(
            "openpyxl is required to read .xlsx files. Run this helper through "
            "the skill's environment: uv run --project "
            "/fluso/user/workspace/skills/security-questionnaire python "
            "scripts/questionnaire.py ..."
        ) from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    grid: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        grid.append([("" if c is None else str(c)) for c in row])
    name = ws.title
    wb.close()
    return grid, name


def load_grid(path: Path, sheet: str | None) -> tuple[list[list[str]], str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        grid, delimiter = read_csv_grid(path)
        return grid, delimiter
    if suffix in (".xlsx", ".xlsm"):
        return read_xlsx_grid(path, sheet)
    raise QError(f"unsupported file type {suffix!r}; use .csv or .xlsx")


# --------------------------------------------------------------------------- #
# Header + column detection.
# --------------------------------------------------------------------------- #

def map_header_row(cells: list[str]) -> dict[str, int]:
    """Map field -> 0-based column index for one candidate header row."""
    mapping: dict[str, int] = {}
    taken: set[int] = set()

    # Pass 1: exact normalized match is authoritative.
    for idx, cell in enumerate(cells):
        norm = normalize(cell)
        if not norm:
            continue
        for field, tokens in HEADER_SYNONYMS.items():
            if field not in mapping and norm in tokens:
                mapping[field] = idx
                taken.add(idx)
                break

    # Pass 2: whole-word fallback for still-unmapped fields.
    for idx, cell in enumerate(cells):
        if idx in taken:
            continue
        words = {normalize(w) for w in re.split(r"[\s/_-]+", str(cell)) if w}
        for field in FIELD_PRIORITY:
            if field in mapping:
                continue
            if words & HEADER_SYNONYMS[field]:
                mapping[field] = idx
                taken.add(idx)
                break
    return mapping


def detect_header(grid: list[list[str]], forced_row: int | None, max_scan: int = 20) -> tuple[int, dict[str, int]]:
    """Return (0-based header row index, field->colindex map)."""
    if forced_row is not None:
        idx = forced_row - 1
        if not 0 <= idx < len(grid):
            raise QError(f"--header-row {forced_row} is out of range (sheet has {len(grid)} rows)")
        return idx, map_header_row(grid[idx])

    best_idx, best_map, best_score = None, {}, 0
    for idx, cells in enumerate(grid[:max_scan]):
        mapping = map_header_row(cells)
        # Reward a usable header: needs a question column plus at least one more.
        score = len(mapping) + (2 if "question" in mapping else 0)
        if score > best_score:
            best_idx, best_map, best_score = idx, mapping, score

    if best_idx is None or "question" not in best_map:
        raise QError(
            "could not locate a header row with a recognizable question column. "
            "Pass --header-row N (1-based) to point at it explicitly."
        )
    return best_idx, best_map


# --------------------------------------------------------------------------- #
# extract
# --------------------------------------------------------------------------- #

def cmd_extract(args: argparse.Namespace) -> int:
    path = Path(args.input)
    if not path.is_file():
        raise QError(f"input file not found: {path}")
    grid, sheet_or_delim = load_grid(path, args.sheet)
    header_idx, mapping = detect_header(grid, args.header_row)

    columns = {
        field: {"index": col, "letter": col_letter(col)}
        for field, col in sorted(mapping.items(), key=lambda kv: kv[1])
    }

    rows = []
    for offset, cells in enumerate(grid[header_idx + 1:], start=header_idx + 2):
        record: dict[str, object] = {"row": offset}
        has_content = False
        for field in ALL_FIELDS:
            if field in mapping:
                idx = mapping[field]
                value = cells[idx].strip() if idx < len(cells) else ""
                record[field] = value
                if field in CONTEXT_FIELDS and value:
                    has_content = True
        # Skip blank spacer rows: no ref/domain/question content at all.
        if has_content:
            rows.append(record)

    result = {
        "source_file": path.name,
        "file_type": path.suffix.lower().lstrip("."),
        "sheet": sheet_or_delim if path.suffix.lower() != ".csv" else None,
        "header_row": header_idx + 1,
        "columns": columns,
        "unmapped_answer_fields": [f for f in ANSWER_FIELDS if f not in mapping],
        "row_count": len(rows),
        "rows": rows,
    }
    write_json(result, args.out)
    if result["unmapped_answer_fields"]:
        print(
            f"note: no column detected for {result['unmapped_answer_fields']}; "
            "fill will skip those fields. Use --header-row if the header was missed.",
            file=sys.stderr,
        )
    return 0


# --------------------------------------------------------------------------- #
# template
# --------------------------------------------------------------------------- #

def cmd_template(args: argparse.Namespace) -> int:
    data = json.loads(Path(args.rows).read_text(encoding="utf-8"))
    rows = data.get("rows", data if isinstance(data, list) else [])
    answers = []
    for record in rows:
        answers.append(
            {
                "row": record.get("row"),
                "ref": record.get("ref", ""),
                "response": "",
                "detail": "",
                "evidence": "",
                "status": "Needs review",
            }
        )
    write_json({"answers": answers}, args.out)
    return 0


# --------------------------------------------------------------------------- #
# fill
# --------------------------------------------------------------------------- #

def load_answers(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    answers = data.get("answers", data) if isinstance(data, dict) else data
    if not isinstance(answers, list):
        raise QError("answers file must be a list, or an object with an 'answers' list")
    return answers


def resolve_target_row(answer, ref_to_row, grid, ref_col, header_row, total_rows):
    """Resolve an answer to a safe target row.

    Returns (row_1based_or_None, warning_or_None). A row is only returned when
    it lands on an existing data row below the header; the sheet is never
    extended, the header is never overwritten, and a row whose 'ref' disagrees
    with the sheet is corrected by ref or skipped, never written blindly.
    """
    ref = str(answer.get("ref", "")).strip()
    raw = answer.get("row")

    if raw is not None:
        try:
            row = int(raw)
        except (TypeError, ValueError):
            return None, f"skipped answer with non-integer row {raw!r}"
        if row <= header_row or row > total_rows:
            return None, (
                f"skipped answer for row {row}: outside the data range "
                f"(rows {header_row + 1}-{total_rows}); the sheet is not extended"
            )
        if ref and ref_col is not None:
            cells = grid[row - 1]
            actual = cells[ref_col].strip() if ref_col < len(cells) else ""
            if actual and actual != ref:
                if ref in ref_to_row:
                    corrected = ref_to_row[ref]
                    return corrected, (
                        f"answer ref {ref!r} did not match row {row} (which is {actual!r}); "
                        f"wrote to row {corrected} located by ref instead"
                    )
                return None, (
                    f"skipped answer: ref {ref!r} does not match row {row} "
                    f"(which is {actual!r}) and no row carries that ref"
                )
        return row, None

    if ref and ref in ref_to_row:
        return ref_to_row[ref], None
    if ref:
        return None, f"skipped answer: ref {ref!r} not found in the sheet"
    return None, f"skipped an answer with no 'row' or 'ref': {answer!r}"


def cmd_fill(args: argparse.Namespace) -> int:
    path = Path(args.input)
    if not path.is_file():
        raise QError(f"input file not found: {path}")
    answers = load_answers(Path(args.answers))

    grid, sheet_or_delim = load_grid(path, args.sheet)
    header_idx, mapping = detect_header(grid, args.header_row)

    ref_to_row: dict[str, int] = {}
    if "ref" in mapping:
        ref_col = mapping["ref"]
        for offset, cells in enumerate(grid[header_idx + 1:], start=header_idx + 2):
            if ref_col < len(cells) and cells[ref_col].strip():
                ref_to_row.setdefault(cells[ref_col].strip(), offset)

    suffix = path.suffix.lower()
    out_path = Path(args.out) if args.out else path.with_name(f"{path.stem}-completed{suffix}")
    if out_path.resolve() == path.resolve():
        raise QError("refusing to overwrite the source questionnaire; choose a different --out")

    writes, warnings = plan_writes(answers, mapping, ref_to_row, grid, header_idx)
    if suffix == ".csv":
        write_csv_out(path, grid, writes, out_path, sheet_or_delim)
    else:
        write_xlsx_out(path, writes, out_path, args.sheet)

    for warning in warnings:
        print(f"warning: {warning}", file=sys.stderr)
    print(
        f"wrote {len(writes)} cell update(s) across {len({r for r, _, _ in writes})} row(s) "
        f"to {out_path}",
        file=sys.stderr,
    )
    return 0


def plan_writes(answers, mapping, ref_to_row, grid, header_idx):
    """Return [(row_1based, col_0based, value), ...] plus collected warnings."""
    writes: list[tuple[int, int, str]] = []
    warnings: list[str] = []
    ref_col = mapping.get("ref")
    header_row = header_idx + 1
    total_rows = len(grid)
    for answer in answers:
        target, warning = resolve_target_row(
            answer, ref_to_row, grid, ref_col, header_row, total_rows
        )
        if warning:
            warnings.append(warning)
        if target is None:
            continue
        for field in ANSWER_FIELDS:
            if field not in answer:
                continue
            if field not in mapping:
                warnings.append(
                    f"row {target}: no '{field}' column in the sheet; value not written"
                )
                continue
            writes.append((target, mapping[field], "" if answer[field] is None else str(answer[field])))
    return writes, warnings


def write_csv_out(path, grid, writes, out_path, delimiter):
    for row1, col0, value in writes:
        idx = row1 - 1
        while idx >= len(grid):
            grid.append([])
        row = grid[idx]
        while col0 >= len(row):
            row.append("")
        row[col0] = value
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        csv.writer(fh, delimiter=delimiter, lineterminator="\n").writerows(grid)


def write_xlsx_out(path, writes, out_path, sheet):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise QError(
            "openpyxl is required to write .xlsx files. Run through the skill's "
            "environment with uv run --project "
            "/fluso/user/workspace/skills/security-questionnaire ..."
        ) from exc
    wb = load_workbook(path)  # keep_vba/styles default: formatting preserved
    ws = wb[sheet] if sheet else wb.active
    for row1, col0, value in writes:
        ws.cell(row=row1, column=col0 + 1, value=value)
    wb.save(out_path)
    wb.close()


# --------------------------------------------------------------------------- #
# shared
# --------------------------------------------------------------------------- #

def write_json(obj: object, out: str | None) -> None:
    text = json.dumps(obj, indent=2, ensure_ascii=False)
    if out:
        Path(out).write_text(text + "\n", encoding="utf-8")
        print(f"wrote {out}", file=sys.stderr)
    else:
        print(text)


def add_sheet_flags(p: argparse.ArgumentParser) -> None:
    """Header/sheet overrides, defined per-subcommand so they work in the
    natural trailing position (e.g. `extract file.xlsx --header-row 3`)."""
    p.add_argument("--header-row", type=int, default=None,
                   help="1-based header row, if auto-detection picks the wrong one")
    p.add_argument("--sheet", default=None, help="worksheet name for .xlsx (default: first)")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="questionnaire.py",
        description="Deterministic read/write for security questionnaire spreadsheets.",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p_extract = sub.add_parser("extract", help="read a questionnaire into JSON rows")
    p_extract.add_argument("input")
    p_extract.add_argument("--out", default=None, help="write JSON here instead of stdout")
    add_sheet_flags(p_extract)
    p_extract.set_defaults(func=cmd_extract)

    p_template = sub.add_parser("template", help="emit a blank answer skeleton from extract JSON")
    p_template.add_argument("rows", help="the JSON produced by 'extract'")
    p_template.add_argument("--out", default=None)
    p_template.set_defaults(func=cmd_template)

    p_fill = sub.add_parser("fill", help="write answers back into a formatting-preserving copy")
    p_fill.add_argument("input")
    p_fill.add_argument("answers", help="JSON with an 'answers' list")
    p_fill.add_argument("--out", default=None, help="output path (default: <input>-completed.<ext>)")
    add_sheet_flags(p_fill)
    p_fill.set_defaults(func=cmd_fill)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        return args.func(args)
    except QError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    except FileNotFoundError as exc:
        print(f"error: file not found: {exc.filename}", file=sys.stderr)
        return 1
    except json.JSONDecodeError as exc:
        print(f"error: invalid JSON ({exc})", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
